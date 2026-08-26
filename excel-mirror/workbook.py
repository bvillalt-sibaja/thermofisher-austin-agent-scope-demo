"""openpyxl-backed model behind the Excel mirror grid.

Grid addressing: 0-indexed (row, col) maps directly to openpyxl's 1-indexed
(row+1, col+1) — grid (0, 0) is cell A1. This is deliberately literal so an
RPA automation using cell references (e.g. "A1", "C5") maps predictably.
"""
import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment

CELL_REF_RE = re.compile(r"^([A-Za-z]+)(\d+)$")

THIN_SIDE = Side(style="thin", color="B7B7B7")
THICK_SIDE = Side(style="thick", color="000000")


def ref_to_rc(ref):
    m = CELL_REF_RE.match(ref.strip())
    if not m:
        raise ValueError(f"Bad cell reference: {ref!r}")
    letters, digits = m.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - 64)
    return int(digits) - 1, col - 1


class WorkbookModel:
    def __init__(self, path, sheet=None):
        self.path = path
        self.filename = os.path.basename(path)
        self.wb = openpyxl.load_workbook(path)
        self.active_sheet = sheet or self.wb.sheetnames[0]
        self._last_find = (-1, -1)

    def sheet_names(self):
        return list(self.wb.sheetnames)

    def switch_sheet(self, name):
        if name in self.wb.sheetnames:
            self.active_sheet = name
            self._last_find = (-1, -1)

    @property
    def ws(self):
        return self.wb[self.active_sheet]

    def get(self, row, col):
        return self.ws.cell(row=row + 1, column=col + 1).value

    def set(self, row, col, value):
        self.ws.cell(row=row + 1, column=col + 1).value = value

    def get_ref(self, ref):
        r, c = ref_to_rc(ref)
        return self.get(r, c)

    def set_ref(self, ref, value):
        r, c = ref_to_rc(ref)
        self.set(r, c, value)

    def find_next(self, text, max_row=200, max_col=30):
        """Row-major search starting just after the last match, wrapping once."""
        if not text:
            return None
        text_l = str(text).lower()
        start_r, start_c = self._last_find
        cells = [(r, c) for r in range(max_row) for c in range(max_col)]
        start_idx = 0
        if start_r >= 0:
            try:
                start_idx = cells.index((start_r, start_c)) + 1
            except ValueError:
                start_idx = 0
        ordered = cells[start_idx:] + cells[:start_idx]
        for r, c in ordered:
            val = self.get(r, c)
            if val is not None and text_l in str(val).lower():
                self._last_find = (r, c)
                return (r, c)
        return None

    def find_row_by_value(self, text, search_col=None):
        """Convenience: return the 0-indexed row whose row contains `text`."""
        pos = self.find_next(text)
        return pos[0] if pos else None

    def save(self, path=None):
        self.wb.save(path or self.path)

    # ---------------------------------------------------------- formatting
    def set_fill(self, row, col, rgb):
        """rgb: 6-hex-char string, e.g. 'FFC000' (no leading '#'). openpyxl needs
        a full 8-char ARGB or the alpha channel defaults to 00 (fully
        transparent) even though it *looks* set — always prepend full alpha."""
        argb = rgb if len(rgb) == 8 else f"FF{rgb}"
        cell = self.ws.cell(row=row + 1, column=col + 1)
        cell.fill = PatternFill(fill_type="solid", fgColor=argb, bgColor=argb)

    def get_fill(self, row, col):
        """Returns the 8-char ARGB fill string (e.g. 'FFFFC000'), or None if no fill."""
        cell = self.ws.cell(row=row + 1, column=col + 1)
        fg = cell.fill.fgColor
        if fg is None or cell.fill.fill_type is None:
            return None
        return fg.rgb

    def set_border(self, row, col, style="thin", sides=None):
        """style: 'thin' or 'thick'. sides: subset of ('top','bottom','left','right'), default all."""
        side = THICK_SIDE if style == "thick" else THIN_SIDE
        sides = sides or ("top", "bottom", "left", "right")
        cell = self.ws.cell(row=row + 1, column=col + 1)
        existing = cell.border
        kwargs = {
            "top": existing.top, "bottom": existing.bottom,
            "left": existing.left, "right": existing.right,
        }
        for s in sides:
            kwargs[s] = side
        cell.border = Border(**kwargs)

    def set_border_range(self, row1, col1, row2, col2, style="thin", outline_only=False):
        """Apply a border across a rectangular range. outline_only=True draws only
        the outer edge (Excel's 'Thick Outside Borders'); otherwise draws all
        cell edges within the range ('All Borders')."""
        for r in range(row1, row2 + 1):
            for c in range(col1, col2 + 1):
                if outline_only:
                    sides = []
                    if r == row1:
                        sides.append("top")
                    if r == row2:
                        sides.append("bottom")
                    if c == col1:
                        sides.append("left")
                    if c == col2:
                        sides.append("right")
                    if sides:
                        self.set_border(r, c, style=style, sides=sides)
                else:
                    self.set_border(r, c, style=style)

    def merge_and_center(self, row1, col1, row2, col2, value=None):
        """Merges the given range and centers its text, Excel's 'Merge & Center'."""
        top_left = self.ws.cell(row=row1 + 1, column=col1 + 1)
        if value is not None:
            top_left.value = value
        self.ws.merge_cells(
            start_row=row1 + 1, start_column=col1 + 1,
            end_row=row2 + 1, end_column=col2 + 1,
        )
        top_left.alignment = Alignment(horizontal="center", vertical="center")

    def merged_ranges(self):
        """Returns a list of (row1, col1, row2, col2) 0-indexed merged ranges."""
        out = []
        for rng in self.ws.merged_cells.ranges:
            out.append((rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1))
        return out

    def has_border(self, row, col):
        b = self.ws.cell(row=row + 1, column=col + 1).border
        return any(s is not None and s.style for s in (b.top, b.bottom, b.left, b.right))
